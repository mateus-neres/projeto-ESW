from loguru import logger
from openpyxl import Workbook, load_workbook
import os
import sys
import tabula
import PyPDF2
import pandas as pd
import numpy as np

# Defina o caminho do JAR
tabula_jar_path = os.path.join(os.path.dirname(sys.argv[0]), 'tabula', 'tabula-1.0.5-jar-with-dependencies.jar')
tabula.environment_info.JAR_PATH = tabula_jar_path

# Definindo o diretório atual
os.chdir(os.path.dirname(__file__))

dir_atual = os.path.dirname(os.path.abspath(sys.argv[0]))

def caminho_absoluto_arquivo(nome_arquivo):
    diretorio_script = os.path.dirname(sys.argv[0])
    caminho_absoluto = os.path.join(diretorio_script, nome_arquivo)
    return caminho_absoluto

# Função para extrair tabelas de um PDF
def pdf_para_tabela(arquivo_pdf):
    logger.info('Iniciando a conversão do PDF para tabelas.')

    
    try:
        tabelas = tabula.read_pdf(caminho_absoluto_arquivo(arquivo_pdf), pages='all', stream=True, multiple_tables=True, encoding='ISO-8859-1')
        
        if not tabelas:
            logger.warning('Nenhuma tabela foi extraída do PDF.')
            return None
        
        logger.info('Conversão do PDF para tabelas concluída com sucesso.')
        return tabelas
    
    except Exception as e:
        logger.error(f'Erro ao converter PDF para tabelas: {e}')
        return None

# Função para salvar tabelas em um arquivo Excel
def tabela_para_excel(tabelas, arquivo_excel):
    logger.info('Iniciando a conversão das tabelas para arquivo Excel.')

    try:
        if not tabelas:
            logger.warning('Nenhuma tabela foi fornecida.')
            return

        # Verifica se o arquivo já existe; se não, cria um novo arquivo Excel
        if os.path.exists(caminho_absoluto_arquivo(arquivo_excel)):
            workbook = load_workbook(caminho_absoluto_arquivo(arquivo_excel))
            sheet = workbook.active
            logger.info('Arquivo Excel existente carregado.')
        else:
            workbook = Workbook()
            sheet = workbook.active
            logger.info('Novo arquivo Excel criado.')

        # Insere as tabelas no Excel
        for tabela in tabelas:
            for linha in tabela.values.tolist():
                if linha:
                    sheet.append(linha)
                else:
                    logger.warning('Linha vazia encontrada e ignorada.')

        workbook.save(caminho_absoluto_arquivo(arquivo_excel))
        logger.info(f'Arquivo Excel salvo com sucesso em: {caminho_absoluto_arquivo(arquivo_excel)}')
    
    except Exception as e:
        logger.error(f'Erro ao salvar a tabela como Excel: {e}')

# Função para extrair texto de um PDF
def ler_pdf(arquivo_pdf):
    logger.info('Iniciando a leitura do PDF.')
   
    try:
        with open(caminho_absoluto_arquivo(arquivo_pdf), 'rb') as arquivo:
            leitor_pdf = PyPDF2.PdfReader(arquivo)
            texto = ''
            
            for pagina in range(len(leitor_pdf.pages)):
                texto += leitor_pdf.pages[pagina].extract_text()
        
        logger.info('Leitura do PDF concluída com sucesso.')
        return texto
    
    except Exception as e:
        logger.error(f'Erro ao ler o PDF: {e}')
        return None

# Função para ajustar valores monetários
def ajustar_valor(valor):
    if pd.isna(valor):
        logger.info('Valor é NaN, retornando None.')
        return None
    
    valor_str = str(valor).strip().replace('.', '').replace(' ', '')
    
    sufixo = ''
    if valor_str.endswith('D') or valor_str.endswith('C'):
        sufixo = valor_str[-1]
        valor_str = valor_str[:-1]
    
    try:
        valor_ajustado = float(valor_str.replace(',', '.'))
        if sufixo == 'D':
            valor_ajustado = -valor_ajustado
        logger.info(f'Valor ajustado: {valor_ajustado}')
        return valor_ajustado
    
    except ValueError:
        logger.error(f'Valor inválido para conversão: {valor_str}')
        return None
