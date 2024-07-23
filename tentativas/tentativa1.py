import tabula
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Função para converter o extrato bancário do Sicredi em PDF para Excel
def pdf_to_excel(pdf_path, excel_path):
    # Usar tabula para extrair tabelas do PDF
    df_list = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True)

    # Criar um novo arquivo Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Preencher o Excel com as tabelas extraídas
    for df in df_list:
        # Converter DataFrame para uma lista de listas
        rows = df.values.tolist()

        # Adicionar linhas à planilha
        for row in rows:
            worksheet.append(row)

    # Destacar colunas importantes
    important_columns = ["Data", "Descrição", "Documento", "Valor (R$)"]
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value in important_columns:
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Destaque com cor dourada

    # Salvar o arquivo Excel
    workbook.save(excel_path)

# Chamar a função para converter o extrato bancário do Sicredi em PDF para Excel
pdf_to_excel("EXTRATOSICREDI.pdf", "extrato_sicredi.xlsx")
